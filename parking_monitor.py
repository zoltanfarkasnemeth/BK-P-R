"""
NAP Hub P+R Parkoló Monitor
============================
Beolvassa a DATEX II statikus és dinamikus adatokat, JSON-ba menti,
majd 5 percenként frissíti az Excel naplót változáskövetéssel.

Statikus feed (kapacitás, pozíció, név):
  https://napphub.kozut.hu/hub-web//datex2/2_3/a0db40f1-6bc2-4059-853a-64eb91a45e9c/getDatex2Data

Dinamikus feed (aktuális foglaltság):
  https://napphub.kozut.hu/hub-web//datex2/2_3/7b21743f-fd9f-43d1-9b5e-77e3eb714c06/getDatex2Data
"""

import requests
import json
import time
import logging
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Konfiguráció
# ---------------------------------------------------------------------------
STATIC_URL  = "https://napphub.kozut.hu/hub-web//datex2/2_3/a0db40f1-6bc2-4059-853a-64eb91a45e9c/getDatex2Data"
DYNAMIC_URL = "https://napphub.kozut.hu/hub-web//datex2/2_3/7b21743f-fd9f-43d1-9b5e-77e3eb714c06/getDatex2Data"

POLL_INTERVAL_SEC = 300          # 5 perc
JSON_FILE   = "parking_data.json"
EXCEL_FILE  = "parking_log.xlsx"
LOG_FILE    = "parking_monitor.log"

# DATEX II névterek
NS = {
    "ns2": "http://datex2.eu/schema/2/2_0",
    "ns3": "http://datex2.eu/schema/2/2_0",
    "ns4": "http://datex2.eu/schema/2/2_0",
    "d2":  "http://datex2.eu/schema/2/2_0",
    # fallback nélkül is keresünk
}

# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# XML segédfüggvények
# ---------------------------------------------------------------------------

def _ns_strip(tag: str) -> str:
    """Visszaadja a tag helyi nevét (névtér nélkül)."""
    return tag.split("}")[-1] if "}" in tag else tag


def _find_text(elem, *local_names) -> str | None:
    """Megkeres egy elemet local névvel bárhol a részfában."""
    for child in elem.iter():
        if _ns_strip(child.tag) in local_names:
            return (child.text or "").strip() or None
    return None


def _find_all(elem, local_name):
    """Visszaadja az összes egyező local-nevű elemet."""
    return [c for c in elem.iter() if _ns_strip(c.tag) == local_name]


def fetch_xml(url: str) -> ET.Element | None:
    """Letölti és parse-olja az XML-t, None ha hiba."""
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        return ET.fromstring(resp.content)
    except Exception as exc:
        log.error("Letöltési hiba (%s): %s", url, exc)
        return None


# ---------------------------------------------------------------------------
# Statikus adatok feldolgozása
# ---------------------------------------------------------------------------

def parse_static(root: ET.Element) -> dict[str, dict]:
    """
    Visszaad egy dict-et  {parking_id -> {...}} formában.
    Kivon: id, nev, kapacitas_osszes, latitude, longitude
    """
    result: dict[str, dict] = {}

    for pr in _find_all(root, "parkingRecord"):
        pid = pr.attrib.get("id", "").strip()
        if not pid:
            continue

        # Név – <value lang="hu"> vagy bármelyik <value>
        nev = None
        for val in _find_all(pr, "value"):
            lang = val.attrib.get("lang", "")
            if lang == "hu" or nev is None:
                nev = (val.text or "").strip()

        # Teljes kapacitás
        cap_txt = _find_text(pr, "parkingNumberOfSpaces", "numberOfSpaces")
        kapacitas = int(cap_txt) if cap_txt and cap_txt.isdigit() else None

        # Koordináták
        lat_txt = _find_text(pr, "latitude")
        lon_txt = _find_text(pr, "longitude")
        lat = float(lat_txt) if lat_txt else None
        lon = float(lon_txt) if lon_txt else None

        result[pid] = {
            "id": pid,
            "nev": nev or pid,
            "kapacitas_osszes": kapacitas,
            "latitude": lat,
            "longitude": lon,
        }

    log.info("Statikus adatok: %d parkoló beolvasva.", len(result))
    return result


# ---------------------------------------------------------------------------
# Dinamikus adatok feldolgozása
# ---------------------------------------------------------------------------

def parse_dynamic(root: ET.Element) -> dict[str, dict]:
    """
    Visszaad egy dict-et {parking_id -> {...}} formában.

    A dinamikus feed szerkezete:
      <parkingRecordStatus xsi:type="ParkingSiteStatus">
        <parkingRecordReference id="P131" version="1"/>
        <parkingStatusOriginTime>...</parkingStatusOriginTime>
        <parkingOccupancy>
          <parkingNumberOfVehicles>500</parkingNumberOfVehicles>
        </parkingOccupancy>
      </parkingRecordStatus>

    A foglalt szám = parkingNumberOfVehicles
    A szabad = kapacitás - foglalt  (majd a merge() számolja)
    """
    result: dict[str, dict] = {}

    for status in _find_all(root, "parkingRecordStatus"):
        # Az ID a <parkingRecordReference id="P131"> attribútumban van
        pid = None
        for ref in _find_all(status, "parkingRecordReference"):
            pid = ref.attrib.get("id", "").strip() or None
            if pid:
                break

        if not pid:
            # Fallback: ha mégis az elem saját attribútumában lenne
            pid = status.attrib.get("id", "").strip() or None

        if not pid:
            continue

        # Foglalt járművek száma
        vehicles_txt = _find_text(status, "parkingNumberOfVehicles")
        # Esetleges alternatív tagek is kezelve
        occupied_txt = _find_text(status, "parkingNumberOfOccupiedSpaces", "numberOfOccupiedSpaces")
        vacant_txt   = _find_text(status, "parkingNumberOfVacantSpaces",   "numberOfVacantSpaces")

        def to_int(val):
            return int(val) if val and val.strip().lstrip("-").isdigit() else None

        foglalt = to_int(vehicles_txt) or to_int(occupied_txt)
        szabad  = to_int(vacant_txt)

        # Időbélyeg
        ts_txt = _find_text(
            status,
            "parkingStatusOriginTime",
            "measurementOrCalculationTime",
            "publicationTime",
        )

        result[pid] = {
            "id":            pid,
            "szabad_helyek": szabad,
            "foglalt_helyek": foglalt,
            "meres_ideje":   ts_txt,
        }

    log.info("Dinamikus adatok: %d parkoló beolvasva.", len(result))
    return result


# ---------------------------------------------------------------------------
# Adatok kombinálása
# ---------------------------------------------------------------------------

def merge(static: dict, dynamic: dict) -> list[dict]:
    """Összekapcsolja a statikus és dinamikus adatokat parking_id alapján."""
    combined = []
    all_ids = sorted(set(static) | set(dynamic))

    for pid in all_ids:
        s = static.get(pid, {})
        d = dynamic.get(pid, {})

        kapacitas = s.get("kapacitas_osszes")
        szabad    = d.get("szabad_helyek")
        foglalt   = d.get("foglalt_helyek")

        # Ha foglalt nincs, de szabad + kapacitás megvan, számítsuk
        if foglalt is None and szabad is not None and kapacitas is not None:
            foglalt = kapacitas - szabad

        # Ha szabad nincs, de foglalt + kapacitás megvan
        if szabad is None and foglalt is not None and kapacitas is not None:
            szabad = kapacitas - foglalt

        combined.append({
            "id":               pid,
            "nev":              s.get("nev", pid),
            "kapacitas_osszes": kapacitas,
            "szabad_helyek":    szabad,
            "foglalt_helyek":   foglalt,
            "latitude":         s.get("latitude"),
            "longitude":        s.get("longitude"),
            "meres_ideje":      d.get("meres_ideje"),
            "lekerdezes_ideje": datetime.now().isoformat(timespec="seconds"),
        })

    return combined


# ---------------------------------------------------------------------------
# JSON mentés
# ---------------------------------------------------------------------------

def save_json(data: list[dict]):
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info("JSON mentve: %s (%d rekord)", JSON_FILE, len(data))


# ---------------------------------------------------------------------------
# Excel napló kezelés
# ---------------------------------------------------------------------------

HEADER = [
    "Időpont",
    "Parkoló ID",
    "Parkoló neve",
    "Kapacitás",
    "Szabad helyek",
    "Foglalt helyek",
    "Mérés ideje",
    "Változás (szabad)",
    "Változás ideje",
    "Latitude",
    "Longitude",
]

COL_WIDTHS = [22, 10, 30, 12, 14, 14, 22, 18, 22, 12, 12]

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CHANGE_FILL   = PatternFill("solid", start_color="FFF2CC")   # sárga – változott
NEG_FILL      = PatternFill("solid", start_color="FCE4D6")   # narancs – csökkent
POS_FILL      = PatternFill("solid", start_color="E2EFDA")   # zöld – nőtt
BORDER_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def _init_excel() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Napló"

    ws.append(HEADER)
    for col_idx, (hdr, width) in enumerate(zip(HEADER, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    return wb


def _load_or_init_excel() -> openpyxl.Workbook:
    if Path(EXCEL_FILE).exists():
        return openpyxl.load_workbook(EXCEL_FILE)
    return _init_excel()


def _get_prev_szabad(ws, parking_id: str) -> int | None:
    """Visszaadja az utolsó rögzített szabad helyek számát az adott parkolóra."""
    last_val = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == parking_id and row[4] is not None:
            last_val = row[4]
    return last_val


def append_to_excel(data: list[dict]):
    wb = _load_or_init_excel()
    ws = wb.active
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for rec in data:
        pid    = rec["id"]
        szabad = rec["szabad_helyek"]

        prev   = _get_prev_szabad(ws, pid)
        valtozas     = None
        valtozas_ideje = None
        row_fill = None

        if szabad is not None and prev is not None and szabad != prev:
            valtozas       = szabad - prev
            valtozas_ideje = now_str
            row_fill = POS_FILL if valtozas > 0 else NEG_FILL

        row = [
            now_str,
            pid,
            rec["nev"],
            rec["kapacitas_osszes"],
            szabad,
            rec["foglalt_helyek"],
            rec["meres_ideje"],
            valtozas,
            valtozas_ideje,
            rec["latitude"],
            rec["longitude"],
        ]
        ws.append(row)

        # Formázás az utolsó sorra
        last_row = ws.max_row
        for col_idx in range(1, len(HEADER) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.border    = BORDER_THIN
            cell.alignment = CENTER
            cell.font      = Font(name="Arial", size=9)
            if row_fill and col_idx == 8:        # Változás oszlop kiemelés
                cell.fill = row_fill
            elif row_fill:
                cell.fill = CHANGE_FILL

        if valtozas is not None:
            log.info(
                "Változás – %s [%s]: %+d (volt: %s, most: %s)",
                rec["nev"], pid, valtozas, prev, szabad,
            )

    wb.save(EXCEL_FILE)
    log.info("Excel frissítve: %s (%d sor)", EXCEL_FILE, ws.max_row - 1)


# ---------------------------------------------------------------------------
# Fő lekérdezési ciklus
# ---------------------------------------------------------------------------

def poll_once(static_cache: dict) -> dict:
    """Egyetlen lekérdezési kör. Visszaadja a (esetleg frissített) statikus cache-t."""
    log.info("=== Lekérdezés indul ===")

    # Statikus adatok: csak az első körben (vagy ha üres a cache)
    if not static_cache:
        root_s = fetch_xml(STATIC_URL)
        if root_s is not None:
            static_cache = parse_static(root_s)
        else:
            log.warning("Statikus feed nem elérhető, üres kapacitásokkal folytatjuk.")

    # Dinamikus adatok
    root_d = fetch_xml(DYNAMIC_URL)
    if root_d is None:
        log.warning("Dinamikus feed nem elérhető, kihagyás.")
        return static_cache

    dynamic = parse_dynamic(root_d)
    combined = merge(static_cache, dynamic)

    save_json(combined)
    append_to_excel(combined)

    return static_cache


def main():
    import sys
    once_mode = "--once" in sys.argv   # GitHub Actions / egyszeri futás

    log.info("NAP Hub P+R parkoló monitor indul.%s", " [EGYSZERI MÓD]" if once_mode else "")
    log.info("Statikus feed : %s", STATIC_URL)
    log.info("Dinamikus feed: %s", DYNAMIC_URL)
    if not once_mode:
        log.info("Lekérdezési intervallum: %d mp", POLL_INTERVAL_SEC)

    static_cache: dict = {}

    if once_mode:
        # GitHub Actions: egyetlen lekérdezés, majd kilép
        try:
            poll_once(static_cache)
        except Exception as exc:
            log.exception("Hiba az egyszeri futásban: %s", exc)
            sys.exit(1)
    else:
        # Helyi folyamatos futás 5 percenként
        while True:
            try:
                static_cache = poll_once(static_cache)
            except Exception as exc:
                log.exception("Váratlan hiba a fő ciklusban: %s", exc)

            log.info("Következő lekérdezés %d mp múlva...", POLL_INTERVAL_SEC)
            time.sleep(POLL_INTERVAL_SEC)


if __name__ == "__main__":
    main()
